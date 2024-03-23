import argparse
import openpyxl
import os

# Get path to root of directory containing images
parser = argparse.ArgumentParser(
    prog='move_images',
    description='Moves images from date folders to sample ID folders')
parser.add_argument('-p', '--path', type=str, required=True)
args = parser.parse_args()
root_path = os.path.abspath(args.path)
print('Using root path:', root_path)
 
# Read the Excel file to get mapping of date folder paths to sample IDs
workbook = openpyxl.load_workbook(os.path.join(root_path, "DateToSampleID.xlsx"))
worksheet = workbook.active
# start with row 1 instead of 0 since we don't care about headers, iterate up to last row in sheet
for row in range(1, worksheet.max_row):
  src_path = worksheet.cell(row = row, column = 1).value
  sample_id = worksheet.cell(row = row, column = 2).value

  print('Source path', src_path, 'Sample ID', sample_id)